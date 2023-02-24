import time
import numpy as np
import pandas as pd
import openpyxl as opx
import openpyxl.utils.dataframe

# import pandas as pd
#
# lst = [123,124,124,124,123,123,123,122,122,122]
# data = pd.DataFrame(lst)
# q = data.std()
# print(data)
# print("--------------------------")
# print(q)

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import time

list = [1,3,2,4,2,56,3,6,3,4,5,7,8,1,2,1,1,3,2]

df = pd.DataFrame(list, columns=["shi"])


wb = Workbook()
ws1 = wb.create_sheet("Sheet_A11")
ws1.title = "Title_A1112"
ws2 = wb.create_sheet("Sheet_B1", 0)
ws2.title = "Title_B"

for row in opx.utils.dataframe.dataframe_to_rows(df, index=True, header=True):
    ws2.append(row)

wb.save(filename ='sample_book.xlsx')
